from dotenv import load_dotenv
from pathlib import Path
import os

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from fastapi import FastAPI, APIRouter, HTTPException, Request, Depends, Response, Query
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Literal
from datetime import datetime, timezone, timedelta
import logging
import uuid
import io
import bcrypt
import jwt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ----------------------------------------------------------------------------
# Setup
# ----------------------------------------------------------------------------
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_SECRET = os.environ['JWT_SECRET']
JWT_ALGORITHM = "HS256"

app = FastAPI(title="Apex Industrial Engineering API")
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


# ----------------------------------------------------------------------------
# Auth helpers
# ----------------------------------------------------------------------------
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(plain: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False


def create_access_token(user_id: str, email: str, role: str) -> str:
    payload = {
        "sub": user_id, "email": email, "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(days=7), "type": "access",
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")


async def get_admin_user(user: dict = Depends(get_current_user)) -> dict:
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return user


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


# ----------------------------------------------------------------------------
# Models
# ----------------------------------------------------------------------------
class RegisterInput(BaseModel):
    name: str
    email: EmailStr
    password: str
    phone: Optional[str] = ""
    company: Optional[str] = ""


class LoginInput(BaseModel):
    email: EmailStr
    password: str


class ProfileUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    company: Optional[str] = None
    address: Optional[str] = None


class ContactInput(BaseModel):
    name: str
    email: EmailStr
    phone: Optional[str] = ""
    subject: Optional[str] = ""
    message: str


class ProposalItem(BaseModel):
    service_slug: str
    title: str
    category: str
    quantity: int = 1


class ProposalInput(BaseModel):
    customer_name: str
    customer_email: EmailStr
    customer_phone: Optional[str] = ""
    company: Optional[str] = ""
    items: List[ProposalItem] = []
    project_type: Optional[str] = "online"  # online / offline
    budget: Optional[str] = ""
    timeline: Optional[str] = ""
    details: Optional[str] = ""


class ProposalUpdate(BaseModel):
    status: Optional[Literal["pending", "in-progress", "delivered", "cancelled"]] = None
    amount: Optional[float] = None
    admin_notes: Optional[str] = None


class ServiceInput(BaseModel):
    title: str
    category: str
    description: str
    image: Optional[str] = ""
    price_from: Optional[float] = 0
    unit: Optional[str] = "project"
    features: List[str] = []
    active: bool = True


# ----------------------------------------------------------------------------
# Auth routes
# ----------------------------------------------------------------------------
def set_auth_cookie(response: Response, token: str):
    response.set_cookie(key="access_token", value=token, httponly=True,
                        secure=True, samesite="none", max_age=604800, path="/")


@api_router.post("/auth/register")
async def register(payload: RegisterInput, response: Response):
    email = payload.email.lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Email already registered")
    user = {
        "id": str(uuid.uuid4()),
        "name": payload.name,
        "email": email,
        "password_hash": hash_password(payload.password),
        "phone": payload.phone or "",
        "company": payload.company or "",
        "address": "",
        "role": "customer",
        "loyalty_points": 0,
        "created_at": now_iso(),
    }
    await db.users.insert_one(user)
    token = create_access_token(user["id"], email, "customer")
    set_auth_cookie(response, token)
    user.pop("password_hash", None)
    user.pop("_id", None)
    return {"access_token": token, "user": user}


@api_router.post("/auth/login")
async def login(payload: LoginInput, response: Response):
    email = payload.email.lower()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(payload.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    token = create_access_token(user["id"], email, user["role"])
    set_auth_cookie(response, token)
    user.pop("password_hash", None)
    user.pop("_id", None)
    return {"access_token": token, "user": user}


@api_router.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    return {"message": "Logged out"}


@api_router.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return user


@api_router.put("/auth/profile")
async def update_profile(payload: ProfileUpdate, user: dict = Depends(get_current_user)):
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    if updates:
        await db.users.update_one({"id": user["id"]}, {"$set": updates})
    updated = await db.users.find_one({"id": user["id"]}, {"_id": 0, "password_hash": 0})
    return updated


# ----------------------------------------------------------------------------
# Public routes
# ----------------------------------------------------------------------------
@api_router.get("/services")
async def list_services(category: Optional[str] = None):
    query = {"active": True}
    if category and category != "all":
        query["category"] = category
    services = await db.services.find(query, {"_id": 0}).to_list(500)
    return services


@api_router.get("/services/categories")
async def list_categories():
    cats = await db.services.distinct("category", {"active": True})
    result = []
    for c in cats:
        sample = await db.services.find_one({"category": c}, {"_id": 0})
        result.append({"category": c, "image": sample.get("image", "") if sample else ""})
    return result


@api_router.post("/contact")
async def create_contact(payload: ContactInput):
    doc = {"id": str(uuid.uuid4()), **payload.model_dump(), "created_at": now_iso(), "handled": False}
    await db.contacts.insert_one(doc)
    doc.pop("_id", None)
    return {"message": "Thank you! We will get back to you shortly.", "id": doc["id"]}


def gen_ref() -> str:
    return "APX-" + datetime.now().strftime("%y%m") + "-" + uuid.uuid4().hex[:5].upper()


@api_router.post("/proposals")
async def create_proposal(payload: ProposalInput, request: Request):
    user_id = None
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if token:
        try:
            decoded = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
            user_id = decoded.get("sub")
        except jwt.InvalidTokenError:
            user_id = None
    doc = {
        "id": str(uuid.uuid4()),
        "ref": gen_ref(),
        "user_id": user_id,
        **payload.model_dump(),
        "status": "pending",
        "amount": 0,
        "admin_notes": "",
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    await db.proposals.insert_one(doc)
    doc.pop("_id", None)
    return {"message": "Proposal submitted successfully", "ref": doc["ref"], "proposal": doc}


# ----------------------------------------------------------------------------
# Customer routes
# ----------------------------------------------------------------------------
@api_router.get("/my/proposals")
async def my_proposals(user: dict = Depends(get_current_user)):
    items = await db.proposals.find({"user_id": user["id"]}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return items


# ----------------------------------------------------------------------------
# Admin routes
# ----------------------------------------------------------------------------
@api_router.get("/admin/proposals")
async def admin_proposals(
    status: Optional[str] = None,
    search: Optional[str] = None,
    admin: dict = Depends(get_admin_user),
):
    query = {}
    if status and status != "all":
        query["status"] = status
    if search:
        query["$or"] = [
            {"customer_name": {"$regex": search, "$options": "i"}},
            {"customer_email": {"$regex": search, "$options": "i"}},
            {"ref": {"$regex": search, "$options": "i"}},
            {"company": {"$regex": search, "$options": "i"}},
        ]
    items = await db.proposals.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return items


@api_router.put("/admin/proposals/{proposal_id}")
async def update_proposal(proposal_id: str, payload: ProposalUpdate, admin: dict = Depends(get_admin_user)):
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    updates["updated_at"] = now_iso()
    result = await db.proposals.update_one({"id": proposal_id}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Proposal not found")
    return await db.proposals.find_one({"id": proposal_id}, {"_id": 0})


@api_router.delete("/admin/proposals/{proposal_id}")
async def delete_proposal(proposal_id: str, admin: dict = Depends(get_admin_user)):
    await db.proposals.delete_one({"id": proposal_id})
    return {"message": "Deleted"}


@api_router.get("/admin/customers")
async def admin_customers(admin: dict = Depends(get_admin_user)):
    users = await db.users.find({"role": "customer"}, {"_id": 0, "password_hash": 0}).to_list(2000)
    result = []
    for u in users:
        props = await db.proposals.find({"user_id": u["id"]}, {"_id": 0}).to_list(1000)
        total_spent = sum(p.get("amount", 0) or 0 for p in props)
        result.append({
            **u,
            "order_count": len(props),
            "total_spent": total_spent,
            "type": "repeat" if len(props) > 1 else "new",
        })
    # also include guest emails from proposals without accounts
    return result


@api_router.get("/admin/contacts")
async def admin_contacts(admin: dict = Depends(get_admin_user)):
    items = await db.contacts.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return items


@api_router.get("/admin/analytics")
async def admin_analytics(admin: dict = Depends(get_admin_user)):
    proposals = await db.proposals.find({}, {"_id": 0}).to_list(5000)
    total_revenue = sum(p.get("amount", 0) or 0 for p in proposals)
    status_counts = {"pending": 0, "in-progress": 0, "delivered": 0, "cancelled": 0}
    service_counter = {}
    monthly = {}
    for p in proposals:
        status_counts[p.get("status", "pending")] = status_counts.get(p.get("status", "pending"), 0) + 1
        for it in p.get("items", []):
            t = it.get("title", "Unknown")
            service_counter[t] = service_counter.get(t, 0) + it.get("quantity", 1)
        created = p.get("created_at", "")[:7]
        if created:
            m = monthly.setdefault(created, {"month": created, "projects": 0, "revenue": 0})
            m["projects"] += 1
            m["revenue"] += p.get("amount", 0) or 0
    top_services = sorted(
        [{"name": k, "count": v} for k, v in service_counter.items()],
        key=lambda x: x["count"], reverse=True
    )[:6]
    customer_count = await db.users.count_documents({"role": "customer"})
    repeat = 0
    for uid in await db.proposals.distinct("user_id"):
        if uid and await db.proposals.count_documents({"user_id": uid}) > 1:
            repeat += 1
    monthly_list = sorted(monthly.values(), key=lambda x: x["month"])
    return {
        "total_revenue": total_revenue,
        "total_projects": len(proposals),
        "total_customers": customer_count,
        "pending_projects": status_counts.get("pending", 0),
        "status_counts": status_counts,
        "top_services": top_services,
        "monthly": monthly_list,
        "repeat_customers": repeat,
        "new_customers": max(customer_count - repeat, 0),
    }


@api_router.get("/admin/export/proposals")
async def export_proposals(admin: dict = Depends(get_admin_user)):
    proposals = await db.proposals.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
    wb = Workbook()
    ws = wb.active
    ws.title = "Projects"
    headers = ["Ref", "Date", "Customer", "Email", "Phone", "Company", "Project Type",
               "Services", "Status", "Amount (INR)", "Budget", "Timeline", "Details"]
    ws.append(headers)
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for p in proposals:
        services = "; ".join(f"{i.get('title')} x{i.get('quantity', 1)}" for i in p.get("items", []))
        ws.append([
            p.get("ref", ""), p.get("created_at", "")[:10], p.get("customer_name", ""),
            p.get("customer_email", ""), p.get("customer_phone", ""), p.get("company", ""),
            p.get("project_type", ""), services, p.get("status", ""), p.get("amount", 0),
            p.get("budget", ""), p.get("timeline", ""), p.get("details", ""),
        ])
    widths = [16, 12, 20, 26, 14, 18, 12, 40, 12, 14, 14, 14, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"apex_projects_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@api_router.get("/admin/services")
async def admin_list_services(admin: dict = Depends(get_admin_user)):
    return await db.services.find({}, {"_id": 0}).to_list(500)


@api_router.post("/admin/services")
async def admin_create_service(payload: ServiceInput, admin: dict = Depends(get_admin_user)):
    slug = payload.title.lower().replace(" ", "-").replace("&", "and")[:60] + "-" + uuid.uuid4().hex[:4]
    doc = {"id": str(uuid.uuid4()), "slug": slug, **payload.model_dump(), "created_at": now_iso()}
    await db.services.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.put("/admin/services/{service_id}")
async def admin_update_service(service_id: str, payload: ServiceInput, admin: dict = Depends(get_admin_user)):
    await db.services.update_one({"id": service_id}, {"$set": payload.model_dump()})
    return await db.services.find_one({"id": service_id}, {"_id": 0})


@api_router.delete("/admin/services/{service_id}")
async def admin_delete_service(service_id: str, admin: dict = Depends(get_admin_user)):
    await db.services.delete_one({"id": service_id})
    return {"message": "Deleted"}


# ----------------------------------------------------------------------------
# Seeding
# ----------------------------------------------------------------------------
IMG_CAD = "https://images.unsplash.com/photo-1581092580497-e0d23cbdf1dc?crop=entropy&cs=srgb&fm=jpg&ixid=M3w3NTY2ODh8MHwxfHNlYXJjaHwxfHxtZWNoYW5pY2FsJTIwQ0FEJTIwZGVzaWdufGVufDB8fHx8MTc4MTg3MDkwOXww&ixlib=rb-4.1.0&q=85"
IMG_DOC = "https://images.unsplash.com/photo-1721244654392-9c912a6eb236?crop=entropy&cs=srgb&fm=jpg&ixid=M3w4NjAzMzJ8MHwxfHNlYXJjaHwyfHxlbmdpbmVlcmluZyUyMHRlY2huaWNhbCUyMGRvY3VtZW50YXRpb24lMjBibHVlcHJpbnR8ZW58MHx8fHwxNzgxODcwOTk3fDA&ixlib=rb-4.1.0&q=85"
IMG_DASH = "https://images.unsplash.com/photo-1551288049-bebda4e38f71?crop=entropy&cs=srgb&fm=jpg&ixid=M3w3NDQ2NDN8MHwxfHNlYXJjaHwxfHxkYXRhJTIwYW5hbHl0aWNzJTIwZGFzaGJvYXJkJTIwc2NyZWVufGVufDB8fHx8MTc4MTg3MDkwOXww&ixlib=rb-4.1.0&q=85"
IMG_PRES = "https://images.pexels.com/photos/6285151/pexels-photo-6285151.jpeg?auto=compress&cs=tinysrgb&dpr=2&h=650&w=940"

SEED_SERVICES = [
    # Mechanical Design & Drafting
    ("2D Drafting", "Mechanical Design & Drafting", "Precise 2D engineering drawings and detailing as per industry standards.", IMG_CAD, 3000, ["AutoCAD", "Industry standards", "Revision control"]),
    ("3D Modeling", "Mechanical Design & Drafting", "Parametric 3D CAD models for equipment, assemblies and layouts.", IMG_CAD, 6000, ["Parametric models", "Assemblies", "Renders"]),
    ("GA Drawings", "Mechanical Design & Drafting", "General Arrangement drawings for plant and equipment.", IMG_CAD, 4500, ["GA layouts", "Dimensioning", "BOM"]),
    ("Fabrication Drawings", "Mechanical Design & Drafting", "Shop-ready fabrication and weldment drawings.", IMG_CAD, 5000, ["Weldments", "Cut lists", "Tolerances"]),
    ("Assembly Drawings", "Mechanical Design & Drafting", "Detailed assembly drawings with exploded views and BOM.", IMG_CAD, 4000, ["Exploded views", "BOM", "Fit checks"]),
    ("Equipment Layouts", "Mechanical Design & Drafting", "Plant equipment layouts optimised for flow and maintenance.", IMG_CAD, 5500, ["Plot plans", "Clash checks", "Access"]),
    ("PDF to AutoCAD 2D Conversion", "Mechanical Design & Drafting", "Convert scanned PDFs and images into editable AutoCAD drawings.", IMG_CAD, 1500, ["Editable DWG", "Scale accurate", "Fast turnaround"]),
    # Engineering Documentation
    ("Equipment Datasheets", "Engineering Documentation", "Technical datasheets for pumps, valves, compressors and more.", IMG_DOC, 2500, ["Pumps & valves", "Compressors", "Vendor-ready"]),
    ("Technical Specifications", "Engineering Documentation", "Detailed technical specifications for equipment and systems.", IMG_DOC, 3000, ["Standards based", "Procurement ready"]),
    ("SOP Preparation", "Engineering Documentation", "Standard Operating Procedures for operations and maintenance.", IMG_DOC, 3500, ["Step-by-step", "Safety notes", "Compliance"]),
    ("Hydro Testing Procedures", "Engineering Documentation", "Pneumatic pressure & hydro testing procedures and reports.", IMG_DOC, 4000, ["Test plans", "Inspection reports", "Compliance"]),
    ("Engineering Calculations (Rotating Equipment)", "Engineering Documentation", "Power, sizing and selection calculations for rotating equipment.", IMG_DOC, 4500, ["Power calc", "Sizing", "Selection"]),
    # Industrial Automation & Reporting
    ("Excel Dashboards", "Industrial Automation & Reporting", "Interactive Excel dashboards for engineering KPIs and reporting.", IMG_DASH, 5000, ["KPIs", "Charts", "Auto refresh"]),
    ("Maintenance Tracking Systems", "Industrial Automation & Reporting", "Preventive maintenance trackers and scheduling systems.", IMG_DASH, 6000, ["PM schedules", "Alerts", "History logs"]),
    ("Internal Workflow Automation", "Industrial Automation & Reporting", "Automate repetitive engineering and reporting workflows.", IMG_DASH, 7000, ["Macros", "AI workflows", "Time saving"]),
    ("Automated Reporting Systems", "Industrial Automation & Reporting", "Auto-generated reports from your operational data.", IMG_DASH, 6500, ["Scheduled", "Templated", "Accurate"]),
    # Digital Technical Services
    ("Advanced Excel Dashboards & Reporting", "Digital Technical Services", "Advanced data dashboards with automation and reporting.", IMG_DASH, 5500, ["Pivot models", "Automation", "Insights"]),
    ("Technical Presentations", "Digital Technical Services", "Professional engineering presentations and proposals.", IMG_PRES, 3000, ["Proposals", "Pitch decks", "Visuals"]),
    ("Website Development", "Digital Technical Services", "Modern websites for engineering and industrial businesses.", IMG_PRES, 25000, ["Responsive", "SEO", "Fast"]),
    ("Digital Marketing", "Digital Technical Services", "Lead generation and digital marketing for industrial firms.", IMG_PRES, 15000, ["SEO", "Campaigns", "Analytics"]),
    ("Image & Video Content Creation", "Digital Technical Services", "Technical image and video content for marketing.", IMG_PRES, 8000, ["Renders", "Videos", "Branding"]),
]


async def seed_services():
    if await db.services.count_documents({}) > 0:
        return
    docs = []
    for title, cat, desc, img, price, feats in SEED_SERVICES:
        slug = title.lower().replace(" ", "-").replace("&", "and").replace("(", "").replace(")", "").replace(",", "")
        docs.append({
            "id": str(uuid.uuid4()), "slug": slug, "title": title, "category": cat,
            "description": desc, "image": img, "price_from": price, "unit": "project",
            "features": feats, "active": True, "created_at": now_iso(),
        })
    await db.services.insert_many(docs)
    logger.info("Seeded %d services", len(docs))


async def seed_admin():
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@example.com").lower()
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    existing = await db.users.find_one({"email": admin_email})
    if existing is None:
        await db.users.insert_one({
            "id": str(uuid.uuid4()), "name": "Apex Admin", "email": admin_email,
            "password_hash": hash_password(admin_password), "phone": "", "company": "Apex Industrial Engineering Solutions",
            "address": "", "role": "admin", "loyalty_points": 0, "created_at": now_iso(),
        })
        logger.info("Seeded admin user")
    elif not verify_password(admin_password, existing["password_hash"]):
        await db.users.update_one({"email": admin_email}, {"$set": {"password_hash": hash_password(admin_password), "role": "admin"}})


@app.on_event("startup")
async def startup():
    await db.users.create_index("email", unique=True)
    await db.users.create_index("id")
    await db.proposals.create_index("user_id")
    await seed_admin()
    await seed_services()


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()


@api_router.get("/")
async def root():
    return {"message": "Apex Industrial Engineering API"}


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)
